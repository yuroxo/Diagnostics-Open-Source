#-*- coding: utf-8 -*-
import time
import openpyxl
import discord
import datetime
from dateutil import tz
import config
import asyncio
import re
import koreanbots
import os
import random
import hcskr
from requests import get

dev = '개발자ㅣRampaka#6441'
pp_list = ['유치원','초등학교','중학교','고등학교','특수학교']
school_list = ['서울특별시','부산광역시','대구광역시','인천광역시','광주광역시','대전광역시','울산광역시','세종특별자치시','경기도','강원도','충청북도','충청남도','전라북도','전라남도','경상북도','경상남도','제주특별자치도']
client = discord.Client()
Bot = koreanbots.Client(client, config.koreanbots_token)

@client.event
async def on_guild_join(guild):
    def check(event):
        return event.target.id == client.user.id
    bot_entry = await guild.audit_logs(action=discord.AuditLogAction.bot_add).find(check)
    embed = discord.Embed(title='진단이', color=0x2F3136)
    embed.add_field(name="안녕하세요? 저는 **진단이** 입니다.", value='\n명령어를 보고 싶으시다면 ``진단아 도움말`` 을 입력하세요.\n\n>>> [**디스코드 지원 서버**](https://discord.gg/XnAqJW2huv)', inline=False)
    embed.set_footer(text=f'{guild.name} 서버에 초대해주셔서 정말로 감사합니다!')
    embed.set_image(url='https://media.discordapp.net/attachments/789461017813712908/800991990568845352/Untitled-1.png')
    try:
        await bot_entry.user.send(embed=embed)
    except:
        pass

@client.event
async def on_error(event, *args, **kwargs):
  pass

@client.event
async def on_connect():
  await client.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f'진단아 도움말'))
  print(f'{client.user.name} (으)로 접속됨.')
  print('─────────────────────────────────────────────────────────────────────────────────')

@client.event
async def my_background_task():
    await client.wait_until_ready()
    while True:
      await client.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f'진단아 도움말'))
      await asyncio.sleep(5)
      try:
        await client.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f'총 {len(client.guilds)}개의 서버에서'))
      except:
        pass
      await asyncio.sleep(5)

@client.event
async def on_message(message):
  if message.content == ('진단아 탈퇴'):
    a = "진단아 승인"
    timeout = 15
    embed = discord.Embed(description=f"회원 탈퇴를 원하시면 15초 이내에 진단아 승인을 입력해 주세요.", colour=0x2F3136)
    embed.set_author(name=f"{message.author.name} 님의 회원탈퇴",icon_url=message.author.avatar_url)
    embed.set_footer(text=dev)
    await message.reply(embed=embed, mention_author=True)

    def check(ylt):
      return ylt.author == message.author and ylt.channel == message.channel
    try:
      ylt = await client.wait_for("message", timeout=timeout, check=check)
    except:
      embed = discord.Embed(description=f"시간을 초과하였습니다, 회원 탈퇴가 거절됩니다.", colour=0x2F3136)
      embed.set_author(name=f"{message.author.name} 님의 회원탈퇴",icon_url=message.author.avatar_url)
      embed.set_footer(text=dev)
      await message.reply(embed=embed, mention_author=True)

    if ylt.content == a:
      file = openpyxl.load_workbook('자가진단.xlsx')
      sheet = file.active
      work = (f'{message.author.id}')
      for i in range(1,1000):
        if str(sheet['A'+str(i)].value) == str(work):
          urn = str(sheet['B' + str(i)].value)
          psw = str(sheet['C' + str(i)].value)
          school = str(sheet['G' + str(i)].value)
          bir = str(sheet['D' + str(i)].value)
          con = str(sheet['E' + str(i)].value)
          level = str(sheet['F' + str(i)].value)
          sheet['A'+str(i)].value = '-'
          sheet['B'+str(i)].value = ''
          sheet['C'+str(i)].value = ''
          sheet['D'+str(i)].value = ''
          sheet['E'+str(i)].value = ''
          sheet['F'+str(i)].value = ''
          sheet['G'+str(i)].value = ''
          break
      try:
        print(f'-ㅣ진단아 회원탈퇴ㅣ{message.author.name} ( {urn} {bir} {con} {school} {level} {psw} )\n')
      except UnboundLocalError:
        embed = discord.Embed(description=f"회원의 정보를 찾을 수가 없습니다, ``진단아 설정`` 으로 설정해 주세요.", colour=0x2F3136)
        embed.set_author(name=f"{message.author.name} 님의 회원탈퇴",icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)
        await message.author.send(embed=embed)
      else:
        file.save('자가진단.xlsx')
        embed = discord.Embed(description=f"회원 탈퇴로 인해, 자가진단 데이터에 저장된 회원정보를 제거합니다.", colour=0x2F3136)
        embed.set_author(name=f"{message.author.name} 님의 회원탈퇴",icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)
        await message.author.send(embed=embed)
    else:
      embed = discord.Embed(description=f"승인을 받지 못하였습니다, 회원 탈퇴가 거절됩니다.", colour=0x2F3136)
      embed.set_author(name=f"{message.author.name} 님의 회원탈퇴",icon_url=message.author.avatar_url)
      embed.set_footer(text=dev)
      await message.reply(embed=embed, mention_author=True)

  if message.content == ('진단아 자가진단'):
    if message.channel.type is discord.ChannelType.private:
      embed = discord.Embed(description=f"자가진단 기능은 디엠 채널에서 사용이 불가능 합니다.", colour=0x2F3136)
      embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
      embed.set_footer(text=dev)
      await message.reply(embed=embed, mention_author=True)
    else:
      text1 = open('이용약관.txt','r')
      text1 = text1.read()
      if str(message.author.id) in str(text1):
        start = time.time()
        file = openpyxl.load_workbook('자가진단.xlsx')
        sheet = file.active
        work = (f'{message.author.id}')
        for i in range(1,1000):
          if str(sheet['A'+str(i)].value) == str(work):
            urn = str(sheet['B' + str(i)].value)
            psw = str(sheet['C' + str(i)].value)
            school = str(sheet['G' + str(i)].value)
            bir = str(sheet['D' + str(i)].value)
            con = str(sheet['E' + str(i)].value)
            level = str(sheet['F' + str(i)].value)
            break
        try:
          data = await hcskr.asyncSelfCheck(urn,bir,con,school,level,psw)
        except UnboundLocalError:
          embed = discord.Embed(description=f"회원의 정보를 찾을 수가 없습니다, ``진단아 설정`` 으로 설정해 주세요.", colour=0x2F3136)
          embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
          embed.set_footer(text=dev)
          await message.reply(embed=embed, mention_author=True)
          pass
        else:
          coddie = data['message']
          print('#ㅣ진단아 자가진단ㅣ' + f'{message.author.name}, {coddie}\n')
          if (data['code']) == 'SUCCESS':
            total = (time.time() - start)
            embed = discord.Embed(description=f"{coddie}\n[ 걸린 시간 : {total} ]", colour=0x2F3136)
            embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

            embed = discord.Embed(title = "자가진단을 완료 하였습니다.", description=f"[ 걸린 시간 : {total} ]\n\n**- 공지\n이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 직접 접속해서 다시 제출하시길 바랍니다.", colour=0x2F3136)
            embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.author.send(embed=embed)
          else:
            embed = discord.Embed(description=f"{coddie}", colour=0x2F3136)
            embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)
      else:
        print('#ㅣ진단아 자가진단ㅣ' + f'{message.author.name}, 이용약관에 동의 하셔야 합니다.\n')
        embed = discord.Embed(description=f"자가진단 서비스를 이용 하실려면 이용약관에 동의 하셔야 합니다.", colour=0x2F3136)
        embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)

  if message.content.startswith('진단아 설정'):     
      text1 = open('이용약관.txt','r')
      text1 = text1.read()
      if str(message.author.id) in str(text1):
        if message.channel.type is discord.ChannelType.private:
          try:
            u = message.content[6:].split(' ')
            file = openpyxl.load_workbook('자가진단.xlsx')
            sheet = file.active
            for i in range(1,1000):
              if sheet['A' + str(i)].value == '-' or sheet['A' + str(i)].value == str(message.author.id):
                sheet['A'+ str(i)].value = str(message.author.id)
                sheet['B'+ str(i)].value = str(u[1])
                if len(u[2]) == 4:
                  if u[2].isdigit() == True:
                    sheet['C'+ str(i)].value = u[2]
                  else:
                    embed = discord.Embed(description=f"비밀번호가 숫자가 아닙니다.", colour=0x2F3136)
                    embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                    embed.set_footer(text=dev)
                    await message.reply(embed=embed, mention_author=True)
                    return
                else:
                  embed = discord.Embed(description=f"비밀번호가 4글자가 아닙니다.", colour=0x2F3136)
                  embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                  embed.set_footer(text=dev)
                  await message.reply(embed=embed, mention_author=True)
                  return

                if len(u[3]) == 6:
                  if u[3].isdigit() == True:
                    sheet['D'+ str(i)].value = u[3]
                  else:
                    embed = discord.Embed(description=f"생년월일이 숫자가 아닙니다.", colour=0x2F3136)
                    embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                    embed.set_footer(text=dev)
                    await message.reply(embed=embed, mention_author=True)
                    return
                else:
                  embed = discord.Embed(description=f"생년월일이 6글자가 아닙니다.", colour=0x2F3136)
                  embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                  embed.set_footer(text=dev)
                  await message.reply(embed=embed, mention_author=True)
                  return
                
                if u[5] in pp_list:
                  sheet['F'+ str(i)].value = u[5]
                else:
                  embed = discord.Embed(description=f"등급을 똑바로 입력 해 주세요.", colour=0x2F3136)
                  embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                  embed.set_footer(text=dev)
                  await message.reply(embed=embed, mention_author=True)
                  return
                  
                if u[4] in school_list:
                  sheet['E'+ str(i)].value = u[4]
                else:
                  embed = discord.Embed(description=f"지역권을 똑바로 입력 해 주세요.", colour=0x2F3136)
                  embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
                  embed.set_footer(text=dev)
                  await message.reply(embed=embed, mention_author=True)
                  return

                sheet['G'+ str(i)].value = u[6]
                break
            
            file.save('자가진단.xlsx')
            embed = discord.Embed(description=f"**- 공지\n이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 직접 접속해서 다시 제출하시길 바랍니다. ", colour=0x2F3136)
            embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)

            print(f'+ㅣ진단아 설정ㅣ{message.author.name} ( {u[1]} {u[3]} {u[4]} {u[6]} {u[5]} {u[2]} )\n')
            embed.add_field(name="이름", value=f"``{u[1]}``", inline=True)
            embed.add_field(name="비밀번호", value=f"``{u[2]}``", inline=True)
            embed.add_field(name="생년원일", value=f"``{u[3]}``", inline=True)
            embed.add_field(name="지역", value=f"``{u[4]}``", inline=True)
            embed.add_field(name="학급", value=f"``{u[5]}``", inline=True)
            embed.add_field(name="학교", value=f"``{u[6]}``", inline=True)
            embed.set_footer(text=dev)
            await message.channel.send(embed=embed,content='https://discord.gg/XnAqJW2huv')
          except:
            embed = discord.Embed(description=f"저장을 하는 중 알수 없는 오류가 발생 하였습니다.", colour=0x2F3136)
            embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)
            pass
        else:
          embed = discord.Embed(description=f"개인정보 설정은 디엠 채널에서만 가능 합니다.", colour=0x2F3136)
          embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
          embed.set_footer(text=dev)
          await message.reply(embed=embed, mention_author=True)
      else:
        embed = discord.Embed(description=f"자가진단 서비스를 이용 하실려면 이용약관에 동의 하셔야 합니다.", colour=0x2F3136)
        embed.set_author(name=f"{message.author.name} 님의 자가진단",icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)


  if message.content == ('진단아 이용약관'): 
    embed = discord.Embed(description=f"* 동의를 거부할 수 있으며, 동의 거부시 서비스 이용에 일부 제한 될 수 있습니다.\n\n**- 공지\n 이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 직접 접속해서 다시 제출하시길 바랍니다. ", colour=0x2F3136)
    embed.set_author(name=f"{message.author.name} 님의 이용약관",icon_url=message.author.avatar_url)
    embed.add_field(name="개인 정보 수집 동의", value=f"``자가진단 봇 에서는 서비스 이용 등 서비스 제공을 위해 아래와 같은 최소한의 개인정보를 수집 하고 있습니다.``", inline= False)
    embed.add_field(name="1. 수집하는 개인정보의 항목", value=f"``이름``,``생년월일``,``학교``,``학급``,``지역``,``자가진단 비밀번호``", inline= False)
    embed.add_field(name="2. 개인정보 수집 방법", value=f"``이용약관 동의 후 명령어 : 진단아 설정 [ 개인정보 ] 입력으로 정보를 수집 합니다.``", inline= False)
    embed.add_field(name="3. 개인정보의 수집 및 이용 목적", value=f"``자동 자가진단 기능을 사용 하기 위해 수집 합니다.``", inline= False)
    embed.add_field(name="4. 개인정보의 보유 및 이용기간", value=f"``자가진단 봇의 서비스 종료일 까지.``", inline= False)
    embed.add_field(name="개인정보 제 3자 제공 안내", value=f"``자가진단 봇 에서는 수집된 정보를 제3자에게 제공하지 않습니다.``", inline= False)
    embed.add_field(name="🔔 교육의 목적으로 개인정보 일부 사용", value=f"학교에서 발표를 목적으로 완전한 개인 정보가 아닌, 지역과 학급 같은 단순 통개 자료를 사용함에 동의 함으로 간주합니다.", inline= False)
    embed.set_footer(text='개발자ㅣRampaka#6441\n이 자가진단 봇 이용으로 일어난 모든 책임은 사용자 에게 있습니다.')
    await message.reply(embed=embed, mention_author=True)

  if message.content == ('진단아 동의'):
    text1 = open('이용약관.txt','r')
    text1 = text1.read()
    if str(message.author.id) in str(text1):
      embed = discord.Embed(description=f'이미 이용약관에 동의 하셨습니다.', colour=0x2F3136)
      embed.set_author(name=f"{message.author.name} 님의 이용약관 서비스",icon_url=message.author.avatar_url)
      embed.set_footer(text=dev)
      await message.reply(embed=embed, mention_author=True)
      pass
    else:
      print('*ㅣ진단아 동의ㅣ' + f'{message.author.name}, 이용약관에 성공적으로 동의 하셨습니다.\n')
      text2 = open('이용약관.txt','a')
      text2.write(str(message.author.id))
      text2.write('\n')
      text2.close()
      embed = discord.Embed(description=f'이용약관에 성공적으로 동의 하셨습니다.', colour=0x2F3136)
      embed.set_author(name=f"{message.author.name} 님의 이용약관 서비스",icon_url=message.author.avatar_url)
      embed.set_footer(text=dev)
      await message.reply(embed=embed, mention_author=True)

  if message.content == ('진단아 도움말'):
    embed = discord.Embed(description=f'>>> [**초대하기**](https://discord.com/api/oauth2/authorize?client_id=789452453401722920&permissions=650368&scope=bot)', colour=0x2F3136)
    embed.add_field(name="진단아 도움말", value=f"``진단이의 정보와 명령어 등을 알려 줍니다.``", inline= False)
    embed.add_field(name="진단아 이용약관", value=f"``진단이를 사용 하려면 수집이 필요한 이용약관을 보여 줍니다.``", inline= False)
    embed.add_field(name="진단아 동의", value=f"``위 명령어의 이용약관을 동의 합니다. 이용약관을 꼭 읽어 주세요. ``", inline= False)
    embed.add_field(name="진단아 설정 [이름] [비밀번호] [생년월일6자] [지역권] [학급] [학교]", value=f"``자동 자가진단 기능에 필요한 정보를 설정 합니다.``", inline= False)
    embed.add_field(name="진단아 탈퇴", value=f"``더 이상 사용하기 싫거나 정보를 삭제해야 할 때 탈퇴를 하면 회원정보가 제거됩니다.``", inline= False)
    embed.add_field(name="진단아 자가진단", value=f"``위 명령어 에서 설정한 정보로 자동 자가진단을 실시 합니다.``", inline= False)
    embed.add_field(name="진단아 양식", value=f"``진단아 설정 에서 더 자세한 양식을 알려 줍니다.``", inline= False)
    embed.add_field(name="진단아 정보", value=f"``진단이의 정보들을 알려 줍니다.``", inline= False)
    embed.set_author(name=f"{message.author.name} 님의 도움말",icon_url=message.author.avatar_url)
    embed.set_image(url='https://media.discordapp.net/attachments/789461017813712908/793512087238082560/999ec44c00beaa90.jpg?width=380&height=71')
    embed.set_footer(text=dev)
    await message.reply(embed=embed,content='https://discord.gg/XnAqJW2huv', mention_author=True)
    
  if message.content == ('진단아 양식'):
    embed = discord.Embed(description=f'**진단아 설정 [이름] [비밀번호] [생년월일6자] [지역권] [학급] [학교]**\n- 예 )진단아 설정 홍길동 0123 030417 서울특별시 고등학교 서울예술고등학교', colour=0x2F3136)
    embed.add_field(name="지역권 [ 아래 보기에서 선택 ]", value=f"```서울특별시\n부산광역시\n대구광역시\n인천광역시\n광주광역시\n대전광역시\n울산광역시\n세종특별자치시\n경기도\n강원도\n충청북도\n충청남도\n전라북도\n전라남도\n경상북도\n경상남도\n제주특별자치도```", inline = True)
    embed.add_field(name="학급 [ 아래 보기에서 선택 ]", value=f"```유치원\n초등학교\n중학교\n고등학교\n특수학교```", inline = True)
    embed.set_author(name=f"{message.author.name} 님의 도움말",icon_url=message.author.avatar_url)
    embed.set_footer(text=dev)
    await message.reply(embed=embed,content='https://discord.gg/XnAqJW2huv', mention_author=True)

  if message.content == ('진단아 정보'):
    embed = discord.Embed(description=f'', colour=0x2F3136)
    embed.set_author(name=f"{message.author.name} 님의 도움말",icon_url=message.author.avatar_url)
    embed.add_field(name="핑", value=f">>> {int((client.latency * 1000))}'ms", inline = False)
    embed.add_field(name="제작 기간", value=f">>> 2020-12-17 ~", inline = False)
    embed.add_field(name="사용 모듈", value=f">>> **[Hcskr](https://github.com/331leo/hcskr_python)** 오픈소스를 이용해 제작된 봇입니다", inline = False)
    embed.add_field(name="프로필 정보", value=f"- 수세밀네님 커미션 \n>>> 인스타그램 : @susamilneh\n디스코드 : Susamilneh#1000\n이메일 : sumichip0215@naver.com ", inline = False)
    embed.set_image(url='https://media.discordapp.net/attachments/789461017813712908/800991990568845352/Untitled-1.png')
    embed.set_footer(text=dev)
    await message.reply(embed=embed,content='https://discord.gg/XnAqJW2huv', mention_author=True)
  
  if message.content.startswith('진단아'):
    BASEURL = "https://api.koreanbots.dev"
    token = config.koreanbots_token
    userID = message.author.id
    response = get(f'{BASEURL}/bots/voted/{userID}', headers={"token":token})
    u = response.json()
    vote = (u['voted'])
    if vote == True:
      pass
    else:
      embed = discord.Embed(title=f'{message.author.name}님은 하트를 누르지 않으셨네요.',description=f'\n>>> **[하트 누르기](https://koreanbots.dev/bots/789452453401722920)** ', colour=0x4F545C)
      embed.set_footer(text='하트를 눌러주시면 진단이의 봇 인기 순위가 올라가요!, 항상 감사합니다.')
      await message.channel.send(embed=embed, mention_author=True)
  
client.loop.create_task(my_background_task())
client.run(config.token)